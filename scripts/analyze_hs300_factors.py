"""
沪深300单因子回测结果汇总分析
从各因子的分层回测数据Excel中提取关键指标，按|IR|降序输出
"""
import openpyxl
import os
import sys

# Fix Windows console encoding
sys.stdout.reconfigure(encoding='utf-8')

OUTPUT_DIR = r"D:\Signal_backtesting\output"

FACTORS = [
    "turnoverRateAvg20d_available",
    "turnoverRateAvg120d_available",
    "ILLIQ_available",
    "TurnoverRateChange_available",
    "PB_available",
    "PCF_available",
    "PE_available",
    "PS_available",
    "OCFPR_available",
    "OCFTD_available",
    "GPG_available",
    "ROA_available",
    "ROE_available",
    "GrossProfitTTM_available",
    "NetProfitTTM_available",
    "RTN1_available",
    "RTN3_available",
    "RTN6_available",
    "RTN12_available",
    "CTA_available",
    "LnFloatCap_available",
    "LnFloatCap1_available",
    "ATER_available",
    "DTA_available",
    "FAT_available",
    "LDTWC_available",
    "TAT_available",
]


def safe_float(v):
    """Safely convert to float."""
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def find_row_by_col_a(ws, keyword, max_row=None):
    """Find row number where column A contains keyword (partial match on raw bytes)."""
    mr = max_row or ws.max_row
    for r in range(1, mr + 1):
        val = ws.cell(r, 1).value
        if val is not None and isinstance(val, str):
            # Handle garbled encoding - encode to latin1 then decode as gbk
            try:
                decoded = val.encode('latin1').decode('gbk', errors='ignore')
            except (UnicodeEncodeError, UnicodeDecodeError):
                decoded = val
            if keyword in decoded or keyword in val:
                return r
    return None


def extract_factor_data(xlsx_path):
    """Extract all relevant data from a single factor's Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]  # First sheet is hs300

    # Build decoded map of column A
    col_a = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is not None and isinstance(val, str):
            try:
                decoded = val.encode('latin1').decode('gbk', errors='ignore')
            except (UnicodeEncodeError, UnicodeDecodeError):
                decoded = val
            col_a[r] = decoded
        elif val is not None:
            col_a[r] = str(val)

    result = {}

    # IC Summary: rows 2-8 are fixed position
    result['rank_ic'] = safe_float(ws.cell(2, 2).value)
    result['ic_std'] = safe_float(ws.cell(3, 2).value)
    result['ir'] = safe_float(ws.cell(4, 2).value)
    result['ic_winrate'] = safe_float(ws.cell(5, 2).value)

    # Find "全区间" row in the excess returns section (扣费前)
    # The header "分层超额收益-扣费前(%)" is around row 18
    # "全区间" row follows the yearly data
    excess_pre_header = None
    for r, txt in col_a.items():
        if '扣费前' in txt and '超额' in txt:
            excess_pre_header = r
            break

    if excess_pre_header:
        # Find 全区间 row after this header
        for r in range(excess_pre_header + 1, excess_pre_header + 10):
            if r in col_a and '全区间' in col_a[r]:
                # Determine how many groups from header row
                n_groups = 0
                for c in range(2, 12):
                    hval = ws.cell(excess_pre_header, c).value
                    if hval is not None and 'group' in str(hval):
                        n_groups += 1

                result['n_groups'] = n_groups
                for g in range(1, 6):
                    col_idx = g + 1
                    val = safe_float(ws.cell(r, col_idx).value)
                    result[f'g{g}'] = val
                break

    # Find Top组合 cumulative/annualized excess
    # Look for 累计超额 pattern
    for r, txt in col_a.items():
        if '累计超额' in txt and '扣费前' in txt:
            result['top_cum_excess'] = safe_float(ws.cell(r, 2).value)
        elif '累计超额' in txt and '扣费后' in txt:
            result['top_cum_excess_net'] = safe_float(ws.cell(r, 2).value)
        elif '年化超额' in txt and '扣费前' in txt:
            result['top_ann_excess'] = safe_float(ws.cell(r, 2).value)
        elif '年化超额' in txt and '扣费后' in txt:
            result['top_ann_excess_net'] = safe_float(ws.cell(r, 2).value)

    # Alternative: look for 累计超额 by searching for specific row patterns
    # Some files use different naming. Let's also search by position.
    # The Top section has: 累计超额(扣费前%), 累计超额(扣费后%), 年化超额(扣费前%), 年化超额(扣费后%)
    for r, txt in col_a.items():
        if '累计' in txt and '超额' in txt:
            if '扣费前' in txt or '扣费后' not in txt:
                if 'top_cum_excess' not in result:
                    result['top_cum_excess'] = safe_float(ws.cell(r, 2).value)
        if '年化' in txt and '超额' in txt and '波动' not in txt:
            if '扣费前' in txt or '扣费后' not in txt:
                if 'top_ann_excess' not in result:
                    result['top_ann_excess'] = safe_float(ws.cell(r, 2).value)

    wb.close()
    return result


# ============================================================
# Main
# ============================================================
results = []

for factor_dir in FACTORS:
    factor_name = factor_dir.replace("_available", "")
    xlsx_path = os.path.join(OUTPUT_DIR, factor_dir, f"{factor_dir}_分层回测数据.xlsx")

    if not os.path.exists(xlsx_path):
        print(f"[跳过] {factor_name}: 文件不存在")
        continue

    try:
        data = extract_factor_data(xlsx_path)
        data['factor'] = factor_name
        results.append(data)
    except Exception as e:
        print(f"[错误] {factor_name}: {e}")
        import traceback; traceback.print_exc()

# Sort by |IR| descending
results.sort(key=lambda x: abs(x.get('ir', 0) or 0), reverse=True)

# ============================================================
# Print Tables
# ============================================================
SEP = "=" * 130
THIN = "-" * 130

print(SEP)
print(f"{'沪深300 单因子回测汇总 (按|IR|降序排列)':^130}")
print(SEP)

# Table 1: IC Summary
print(f"\n{'一、IC指标汇总':^130}")
print(THIN)
print(f"{'#':>3} {'因子':<26} {'Rank IC均值':>12} {'IC标准差':>10} {'IR':>10} {'IC胜率%':>10} {'|IR|':>10}")
print(THIN)
for i, r in enumerate(results, 1):
    ic = f"{r['rank_ic']:.4f}" if r.get('rank_ic') is not None else "N/A"
    std = f"{r['ic_std']:.4f}" if r.get('ic_std') is not None else "N/A"
    ir = f"{r['ir']:.4f}" if r.get('ir') is not None else "N/A"
    wr = f"{r['ic_winrate']:.2f}" if r.get('ic_winrate') is not None else "N/A"
    abs_ir = f"{abs(r['ir']):.4f}" if r.get('ir') is not None else "N/A"
    print(f"{i:>3} {r['factor']:<26} {ic:>12} {std:>10} {ir:>10} {wr:>10} {abs_ir:>10}")

# Table 2: Group Excess Returns
print(f"\n{'二、分层超额收益 - 扣费前(%) - 全区间':^130}")
print(THIN)
print(f"{'#':>3} {'因子':<26} {'G1':>8} {'G2':>8} {'G3':>8} {'G4':>8} {'G5':>8} {'G5-G1':>10} {'分组数':>6} {'单调性':>8}")
print(THIN)
for i, r in enumerate(results, 1):
    n_groups = r.get('n_groups', 5)
    gs = []
    for g in range(1, 6):
        val = r.get(f'g{g}')
        gs.append(f"{val:.2f}" if val is not None else "  --")

    # Compute spread using last group vs first
    g_vals = [r.get(f'g{g}') for g in range(1, n_groups + 1)]
    if g_vals and g_vals[0] is not None and g_vals[-1] is not None:
        spread = f"{g_vals[-1] - g_vals[0]:.2f}"
    else:
        spread = "N/A"

    # Monotonicity check
    valid = [v for v in g_vals if v is not None]
    if len(valid) >= 3:
        inc = all(valid[j] <= valid[j+1] for j in range(len(valid)-1))
        dec = all(valid[j] >= valid[j+1] for j in range(len(valid)-1))
        if inc:
            mono = "递增"
        elif dec:
            mono = "递减"
        else:
            mono = "非单调"
    else:
        mono = "N/A"

    print(f"{i:>3} {r['factor']:<26} {gs[0]:>8} {gs[1]:>8} {gs[2]:>8} {gs[3]:>8} {gs[4]:>8} {spread:>10} {n_groups:>6} {mono:>8}")

# Table 3: Top Portfolio
print(f"\n{'三、Top组合表现汇总':^130}")
print(THIN)
print(f"{'#':>3} {'因子':<26} {'累计超额%(扣费前)':>18} {'年化超额%(扣费前)':>18} {'累计超额%(扣费后)':>18} {'年化超额%(扣费后)':>18}")
print(THIN)
for i, r in enumerate(results, 1):
    tc = f"{r['top_cum_excess']:.2f}" if r.get('top_cum_excess') is not None else "N/A"
    ta = f"{r['top_ann_excess']:.2f}" if r.get('top_ann_excess') is not None else "N/A"
    tcn = f"{r['top_cum_excess_net']:.2f}" if r.get('top_cum_excess_net') is not None else "N/A"
    tan = f"{r['top_ann_excess_net']:.2f}" if r.get('top_ann_excess_net') is not None else "N/A"
    print(f"{i:>3} {r['factor']:<26} {tc:>18} {ta:>18} {tcn:>18} {tan:>18}")

# Table 4: Classification
print(f"\n{'四、因子质量分类':^130}")
print(THIN)

strong = [r for r in results if r.get('ir') is not None and abs(r['ir']) >= 0.3]
medium = [r for r in results if r.get('ir') is not None and 0.15 <= abs(r['ir']) < 0.3]
weak = [r for r in results if r.get('ir') is not None and abs(r['ir']) < 0.15]

print(f"\n[强因子] |IR| >= 0.3  共 {len(strong)} 个")
for r in strong:
    d = "正向(值大为优)" if r['ir'] > 0 else "反向(值小为优)"
    print(f"    {r['factor']:<26} IR={r['ir']:.4f}  {d}")

print(f"\n[中等因子] 0.15 <= |IR| < 0.3  共 {len(medium)} 个")
for r in medium:
    d = "正向(值大为优)" if r['ir'] > 0 else "反向(值小为优)"
    print(f"    {r['factor']:<26} IR={r['ir']:.4f}  {d}")

print(f"\n[弱因子] |IR| < 0.15  共 {len(weak)} 个")
for r in weak:
    d = "正向(值大为优)" if r['ir'] > 0 else "反向(值小为优)"
    print(f"    {r['factor']:<26} IR={r['ir']:.4f}  {d}")

print(f"\n共计: {len(results)} 个因子")
print(SEP)
